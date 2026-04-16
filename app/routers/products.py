from fastapi import APIRouter, Depends, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from typing import Optional, List
from io import BytesIO
from openpyxl import Workbook, load_workbook
from .. import models, schemas, auth
from ..database import get_db
from ..config import settings

router = APIRouter(prefix="/products", tags=["products"])
templates = Jinja2Templates(directory=settings.TEMPLATES_DIR)


@router.get("", response_class=HTMLResponse)
async def list_products(
    request: Request,
    category_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_login)
):
    query = db.query(models.Product).join(models.Category)

    if category_id:
        query = query.filter(models.Product.category_id == category_id)

    products = query.order_by(models.Product.name).all()
    categories = db.query(models.Category).order_by(models.Category.name).all()

    return templates.TemplateResponse(
        request, "products/list.html", {"products": products,
            "categories": categories,
            "selected_category": category_id,
            "user": user,
            "is_admin": user.role == "admin"
        }
    )


@router.get("/add", response_class=HTMLResponse)
async def add_product_page(
    request: Request,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    categories = db.query(models.Category).order_by(models.Category.name).all()
    return templates.TemplateResponse(
        request, "products/form.html", {"categories": categories, "user": user, "product": None}
    )


@router.post("/add")
async def add_product(
    request: Request,
    name: str = Form(...),
    category_id: int = Form(...),
    cost_price: float = Form(...),
    selling_price: float = Form(...),
    store_quantity: float = Form(0),
    shop_quantity: float = Form(0),
    reorder_level: int = Form(10),
    unit: str = Form("piece"),
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    # Get form data for pack categories
    form_data = await request.form()
    pack_names = form_data.getlist("pack_names[]")
    pack_sizes = form_data.getlist("pack_sizes[]")
    pack_prices = form_data.getlist("pack_prices[]")

    product = models.Product(
        name=name,
        category_id=category_id,
        cost_price=cost_price,
        selling_price=selling_price,
        store_quantity=store_quantity,
        shop_quantity=shop_quantity,
        reorder_level=reorder_level,
        unit=unit
    )
    db.add(product)
    db.flush()  # Get the product ID

    # Add pack categories
    for i in range(len(pack_names)):
        if pack_names[i] and pack_sizes[i] and pack_prices[i]:
            pack = models.ProductPack(
                product_id=product.id,
                name=pack_names[i],
                pack_size=int(pack_sizes[i]),
                pack_price=float(pack_prices[i])
            )
            db.add(pack)

    db.commit()
    return RedirectResponse(url="/products", status_code=302)


@router.get("/{product_id}/edit", response_class=HTMLResponse)
async def edit_product_page(
    request: Request,
    product_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    categories = db.query(models.Category).order_by(models.Category.name).all()
    return templates.TemplateResponse(
        request, "products/form.html", {"categories": categories, "user": user, "product": product}
    )


@router.post("/{product_id}/edit")
async def edit_product(
    request: Request,
    product_id: int,
    name: str = Form(...),
    category_id: int = Form(...),
    cost_price: float = Form(...),
    selling_price: float = Form(...),
    store_quantity: float = Form(0),
    shop_quantity: float = Form(0),
    reorder_level: int = Form(10),
    unit: str = Form("piece"),
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    # Get form data for pack categories
    form_data = await request.form()
    pack_ids = form_data.getlist("pack_ids[]")
    pack_names = form_data.getlist("pack_names[]")
    pack_sizes = form_data.getlist("pack_sizes[]")
    pack_prices = form_data.getlist("pack_prices[]")

    product.name = name
    product.category_id = category_id
    product.cost_price = cost_price
    product.selling_price = selling_price
    product.store_quantity = store_quantity
    product.shop_quantity = shop_quantity
    product.reorder_level = reorder_level
    product.unit = unit

    # Track which pack IDs are in the form (to delete removed ones)
    submitted_pack_ids = set()

    # Update or create pack categories
    for i in range(len(pack_names)):
        if pack_names[i] and pack_sizes[i] and pack_prices[i]:
            pack_id = pack_ids[i] if i < len(pack_ids) and pack_ids[i] else None

            if pack_id:
                # Update existing pack
                pack_id = int(pack_id)
                submitted_pack_ids.add(pack_id)
                pack = db.query(models.ProductPack).filter(models.ProductPack.id == pack_id).first()
                if pack:
                    pack.name = pack_names[i]
                    pack.pack_size = int(pack_sizes[i])
                    pack.pack_price = float(pack_prices[i])
            else:
                # Create new pack
                pack = models.ProductPack(
                    product_id=product.id,
                    name=pack_names[i],
                    pack_size=int(pack_sizes[i]),
                    pack_price=float(pack_prices[i])
                )
                db.add(pack)

    # Delete packs that were removed from the form
    for existing_pack in product.packs:
        if existing_pack.id not in submitted_pack_ids:
            db.delete(existing_pack)

    db.commit()

    return RedirectResponse(url="/products", status_code=302)


@router.post("/{product_id}/delete")
async def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(product)
    db.commit()
    return RedirectResponse(url="/products", status_code=302)


# Excel Upload/Download
@router.get("/upload", response_class=HTMLResponse)
async def upload_page(
    request: Request,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    categories = db.query(models.Category).order_by(models.Category.name).all()
    return templates.TemplateResponse(
        request, "products/upload.html", {"user": user, "categories": categories}
    )


@router.get("/template")
async def download_template(
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    """Download Excel template for product upload"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = ["name", "category", "cost_price", "selling_price", "store_quantity",
               "shop_quantity", "reorder_level", "unit", "pack_size", "pack_price"]
    ws.append(headers)

    # Add column descriptions in row 2
    descriptions = [
        "Product name (required)",
        "Category name (required, must exist)",
        "Cost per unit (required)",
        "Selling price per unit (required)",
        "Warehouse stock (default: 0)",
        "Shop stock (default: 0)",
        "Low stock alert level (default: 10)",
        "Unit type: piece/bottle/pack/kg/carton/box",
        "Units per pack (for pack sales, default: 1)",
        "Pack price (optional, for bulk discount)"
    ]
    ws.append(descriptions)

    # Sample data rows
    samples = [
        ["Coca-Cola 50cl", "Soft Drinks", 100, 150, 50, 20, 20, "bottle", 12, 1500],
        ["Ice Cream Vanilla 1L", "Frozen Foods", 150, 250, 10, 5, 5, "pack", 1, ""],
        ["Frozen Chicken 1kg", "Frozen Foods", 800, 1200, 20, 10, 5, "pack", 1, ""],
    ]
    for sample in samples:
        ws.append(sample)

    # Style headers
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    # Style descriptions
    for col in range(1, len(descriptions) + 1):
        ws.cell(row=2, column=col).font = ws.cell(row=2, column=col).font.copy(italic=True)

    # Add Categories sheet
    ws_cat = wb.create_sheet("Categories")
    ws_cat.append(["Available Categories"])
    categories = db.query(models.Category).order_by(models.Category.name).all()
    for cat in categories:
        ws_cat.append([cat.name])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_upload_template.xlsx"}
    )


@router.post("/upload")
async def upload_products(
    request: Request,
    file: UploadFile = File(...),
    update_existing: bool = Form(False),
    stock_mode: str = Form("replace"),  # "replace" or "add"
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    """Upload products from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        return templates.TemplateResponse(
        request, "products/upload.html", {"user": user, "error": "Please upload an Excel file (.xlsx or .xls)",
             "categories": db.query(models.Category).all()}
        )

    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active

        # Get categories map
        categories = {c.name.lower(): c for c in db.query(models.Category).all()}

        added = 0
        updated = 0
        errors = []

        # Skip header row(s) - start from row 3 if row 2 looks like descriptions
        start_row = 2
        if ws.cell(row=2, column=1).value and "required" in str(ws.cell(row=2, column=1).value).lower():
            start_row = 3

        for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            # Skip empty rows
            if not row[0]:
                continue

            try:
                name = str(row[0]).strip()
                category_name = str(row[1]).strip() if row[1] else None
                cost_price = float(row[2]) if row[2] else 0
                selling_price = float(row[3]) if row[3] else 0
                store_quantity = float(row[4]) if row[4] else 0
                shop_quantity = float(row[5]) if row[5] else 0
                reorder_level = int(row[6]) if row[6] else 10
                unit = str(row[7]).strip() if row[7] else "piece"
                pack_size = int(row[8]) if row[8] else 1
                pack_price = float(row[9]) if row[9] else None

                # Validate category
                if not category_name:
                    errors.append(f"Row {row_num}: Category is required")
                    continue

                category = categories.get(category_name.lower())
                if not category:
                    errors.append(f"Row {row_num}: Category '{category_name}' not found")
                    continue

                # Check if product exists
                existing = db.query(models.Product).filter(
                    models.Product.name == name
                ).first()

                if existing:
                    if update_existing:
                        existing.category_id = category.id
                        existing.cost_price = cost_price
                        existing.selling_price = selling_price

                        # Handle stock based on mode
                        if stock_mode == "add":
                            existing.store_quantity = (existing.store_quantity or 0) + store_quantity
                            existing.shop_quantity = (existing.shop_quantity or 0) + shop_quantity
                        else:  # replace
                            existing.store_quantity = store_quantity
                            existing.shop_quantity = shop_quantity

                        existing.reorder_level = reorder_level
                        existing.unit = unit
                        existing.pack_size = pack_size if pack_size > 1 else 1
                        existing.pack_price = pack_price if pack_size > 1 and pack_price else None
                        updated += 1
                    else:
                        errors.append(f"Row {row_num}: Product '{name}' already exists")
                        continue
                else:
                    product = models.Product(
                        name=name,
                        category_id=category.id,
                        cost_price=cost_price,
                        selling_price=selling_price,
                        store_quantity=store_quantity,
                        shop_quantity=shop_quantity,
                        reorder_level=reorder_level,
                        unit=unit,
                        pack_size=pack_size if pack_size > 1 else 1,
                        pack_price=pack_price if pack_size > 1 and pack_price else None
                    )
                    db.add(product)
                    added += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.commit()

        return templates.TemplateResponse(
        request, "products/upload.html", {"user": user,
                "categories": db.query(models.Category).all(),
                "success": f"Added {added} products, updated {updated} products",
                "errors": errors if errors else None
            }
        )

    except Exception as e:
        return templates.TemplateResponse(
        request, "products/upload.html", {"user": user, "error": f"Error processing file: {str(e)}",
             "categories": db.query(models.Category).all()}
        )


@router.get("/export")
async def export_products(
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    """Export all products to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = ["name", "category", "cost_price", "selling_price", "store_quantity",
               "shop_quantity", "reorder_level", "unit", "pack_size", "pack_price"]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    # Add products
    products = db.query(models.Product).join(models.Category).order_by(models.Product.name).all()
    for p in products:
        ws.append([
            p.name,
            p.category.name,
            p.cost_price,
            p.selling_price,
            p.store_quantity,
            p.shop_quantity,
            p.reorder_level,
            p.unit,
            p.pack_size or 1,
            p.pack_price or ""
        ])

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"}
    )


# Categories
@router.get("/categories", response_class=HTMLResponse)
async def list_categories(
    request: Request,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_login)
):
    categories = db.query(models.Category).order_by(models.Category.name).all()
    return templates.TemplateResponse(
        request, "products/categories.html", {"categories": categories, "user": user, "is_admin": user.role == "admin"}
    )


@router.post("/categories/add")
async def add_category(
    name: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    category = models.Category(name=name, description=description)
    db.add(category)
    db.commit()
    return RedirectResponse(url="/products/categories", status_code=302)


@router.post("/categories/{category_id}/delete")
async def delete_category(
    category_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(auth.require_admin)
):
    category = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    # Check if category has products
    product_count = db.query(models.Product).filter(models.Product.category_id == category_id).count()
    if product_count > 0:
        raise HTTPException(status_code=400, detail="Cannot delete category with products")

    db.delete(category)
    db.commit()
    return RedirectResponse(url="/products/categories", status_code=302)
